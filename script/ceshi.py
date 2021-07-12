from openpyxl import Workbook
from chinese_calendar import is_workday
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import pandas
import datetime
import calendar
import winreg


def get_desktop():
	key = winreg.OpenKey(
		winreg.HKEY_CURRENT_USER,
		r'Software\Microsoft\Windows\CurrentVersion\Explorer\Shell Folders'
	)
	return winreg.QueryValueEx(key, 'Desktop')[0]


def excel(data, name):

	item = {
		'正常': 10,
		'迟到': 100,
		'早退': 100,
		'旷工': 1000
	}

	date = data[0].check_date
	month_day = calendar.monthrange(date.year, date.month)
	month_first = date.replace(day=1)
	month_end = date.replace(day=month_day[1])
	month_date = pandas.bdate_range(month_first, month_end)
	whole_month = []
	for o in month_date:
		if is_workday(o):
			whole_month.append(o.date())

	wb = Workbook()
	ws = wb.active
	ws.merge_cells('A1:J1')
	ws['A1'] = '%s-%s月考勤' % (whole_month[0].year, whole_month[0].month)
	ws.append(['姓名', '事假', '病假', '婚假', '丧假', '迟到 早退', '旷工', '应出勤', '实际出勤', '备注'])

	for i in name:
		c = data.filter(check_name=i.user_name)
		late = 0
		absenteeism = 0
		remarks = ''
		types = {'事假': 0, '病假': 0, '婚假': 0, '丧假': 0}
		absenteeism += len(whole_month) - len(c)
		for a in c:
			# print(a.check_num, a.check_name, a.check_date)
			plus = item[a.check_morning] + item[a.check_noon] + item[a.check_afternoon]
			if a.check_state in types:
				types[a.check_state] += float(a.check_day)
			if plus >= 2000:
				absenteeism += 1
				remarks += '%s旷工一天，' % a.check_date
			elif 100 <= plus <= 300:
				if plus == 120:
					late += 1
					if a.check_afternoon == '早退':
						remarks += '%s早退，' % a.check_date
					else:
						remarks += '%s迟到，' % a.check_date
				elif plus == 210:
					late += 2
					if a.check_afternoon == '早退':
						remarks += '%s迟到及早退，' % a.check_date
					else:
						remarks += '%s迟到2次，' % a.check_date
				elif plus == 300:
					late += 3
					remarks += '%s迟到2次及早退，' % a.check_date
			elif 1000 <= plus <= 1300:
				if plus == 1020:
					absenteeism += 0.5
					remarks += '%s旷工半天，' % a.check_date
				elif plus == 1110:
					absenteeism += 0.5
					late += 1
					if a.check_afternoon == '早退':
						remarks += '%s早退及旷工，' % a.check_date
					else:
						remarks += '%s迟到及旷工，' % a.check_date
				elif plus == 1200:
					absenteeism += 0.5
					late += 2
					if a.check_afternoon == '早退':
						remarks += '%s迟到、早退及旷工，' % a.check_date
					else:
						remarks += '%s迟到2次及旷工，' % a.check_date
		actual_attendance = len(whole_month) - absenteeism - types['事假'] - types['病假'] - types['婚假'] - types['丧假']
		if late == 0:
			late = None
		if absenteeism == 0:
			absenteeism = None
		if types['事假'] == 0:
			types['事假'] = None
		if types['病假'] == 0:
			types['病假'] = None
		if types['婚假'] == 0:
			types['婚假'] = None
		if types['丧假'] == 0:
			types['丧假'] = None

		ws.append([
			i.user_name,
			types['事假'],
			types['病假'],
			types['婚假'],
			types['丧假'],
			late,
			absenteeism,
			len(whole_month),
			actual_attendance,
			remarks
		])

	# max_column = ws.max_column
	max_row = ws.max_row
	al = Alignment(horizontal='center', vertical='center')
	al1 = Alignment(horizontal='right', vertical='center')
	font1 = Font(size=14, bold=True, name='微软雅黑')
	ws['A1'].font = font1
	ws.row_dimensions[1].height = 26
	ws.column_dimensions['J'].width = 36
	for row in ws['A1:J%s' % max_row]:
		for cell in row:
			cell.alignment = al
	for row1 in ws['J2:J%s' % max_row]:
		for cell1 in row1:
			cell1.alignment = al1

	b = r'%s\%s-%s月汇总.xlsx' % (get_desktop(), date.year, date.month)
	wb.save(b)
